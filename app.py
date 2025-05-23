from flask import Flask, request, jsonify
from flask_cors import CORS
from flask_migrate import Migrate
from models import db, Award, Participant, Result
import random
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
import os
from sqlalchemy import select, func, or_, and_
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import re
from datetime import datetime

app = Flask(__name__)
# Cấu hình CORS chi tiết hơn
CORS(app, resources={
    r"/api/*": {
        "origins": "*",  # Cho phép tất cả các origin trong development
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization"],
        "supports_credentials": True
    }
})

app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB limit

db.init_app(app)
migrate = Migrate(app, db)

# Định nghĩa thư mục để lưu file tạm thời
UPLOAD_FOLDER = 'temp_uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route('/api/awards', methods=['GET'])
def get_awards():
    awards = Award.query.all()
    return jsonify([{
        'id': award.id,
        'ten_giai': award.ten_giai,
        'loai_phuong_thuc': award.loai_phuong_thuc,
        'gia_tri': award.gia_tri,
        'da_duoc_phat': award.da_duoc_phat
    } for award in awards])

@app.route('/api/quay-so', methods=['POST'])
def quay_so():
    data = request.get_json()
    # Kiểm tra cả award_id và awardId
    award_id = data.get('award_id') or data.get('awardId')
    check_only = data.get('check_only', False)

    if not award_id:
        return jsonify({'error': 'Thiếu thông tin giải thưởng'}), 400

    try:
        # Lấy một người chơi ngẫu nhiên chưa trúng giải
        potential_winner = db.session.query(Participant)\
            .outerjoin(Result, Participant.id == Result.participant_id)\
            .filter(Result.id == None)\
            .order_by(func.random())\
            .first()

        if not potential_winner:
            return jsonify({'error': 'Không còn người chơi nào để quay số'}), 400

        # Kiểm tra xem người chơi đã trúng giải nào chưa
        existing_win = db.session.query(Result)\
            .join(Participant)\
            .filter(
                or_(
                    Participant.id == potential_winner.id,
                    and_(
                        Participant.ho_ten == potential_winner.ho_ten,
                        Participant.dia_chi == potential_winner.dia_chi
                    )
                )
            ).first()

        if existing_win:
            # Nếu đã trúng, trả về thông tin người chơi và giải đã trúng
            award = db.session.query(Award).filter_by(id=existing_win.award_id).first()
            return jsonify({
                'error': f'Đại lý đã trúng {award.ten_giai}',
                'participant': {
                    'id': potential_winner.id,
                    'ho_ten': potential_winner.ho_ten,
                    'dia_chi': potential_winner.dia_chi
                },
                'existing_award': {
                    'id': award.id,
                    'ten_giai': award.ten_giai
                }
            }), 409  # Conflict status code

        # Nếu chỉ kiểm tra (check_only=True), trả về thông tin người chơi mà không lưu vào DB
        if check_only:
            return jsonify({
                'participant': {
                    'id': potential_winner.id,
                    'ho_ten': potential_winner.ho_ten,
                    'dia_chi': potential_winner.dia_chi
                }
            })

        # Nếu không phải check_only, tạo kết quả mới và lưu vào DB
        new_result = Result(
            participant_id=potential_winner.id,
            award_id=award_id,
            thoi_gian=datetime.utcnow()
        )
        db.session.add(new_result)
        db.session.commit()

        return jsonify({
            'participant': {
                'id': potential_winner.id,
                'ho_ten': potential_winner.ho_ten,
                'dia_chi': potential_winner.dia_chi
            },
            'award': {
                'id': award_id,
                'ten_giai': db.session.query(Award).filter_by(id=award_id).first().ten_giai
            }
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/results', methods=['GET'])
def get_results():
    search_term = request.args.get('search', '')
    filter_award = request.args.get('award', '')

    query = Result.query.join(Participant).join(Award)

    if search_term:
        query = query.filter(or_(
            Participant.ho_ten.ilike(f'%{search_term}%'),
            Participant.dia_chi.ilike(f'%{search_term}%'),
            Participant.id.ilike(f'%{search_term}%'),
            Participant.ma_vung.ilike(f'%{search_term}%')
        ))

    if filter_award:
        query = query.filter(Award.id == filter_award)

    results = query.order_by(Result.thoi_gian.desc()).all()
    
    return jsonify({
        'results': [{
            'id': result.id,
            'participant': {
                'id': result.participant.id,
                'ho_ten': result.participant.ho_ten,
                'dia_chi': result.participant.dia_chi,
                'ma_vung': result.participant.ma_vung
            },
            'award': {
                'id': result.award.id,
                'ten_giai': result.award.ten_giai,
                'gia_tri': result.award.gia_tri
            },
            'thoi_gian': result.thoi_gian.isoformat()
        } for result in results],
        'total': len(results)
    })

@app.route('/api/participants/upload', methods=['POST'])
def upload_participants():
    app.logger.info("upload_participants function called")
    if 'file' not in request.files:
        return jsonify({'error': 'Không có file được tải lên'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Không có file được chọn'}), 400
    
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Chỉ chấp nhận file Excel (.xlsx)'}), 400
    
    try:
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        with open(filepath, 'rb') as f:
            # Use the read-only mode and data_only=True to read cell values
            wb = load_workbook(filename=f, read_only=True, data_only=True)
            sheet = wb.active

            # Initialize counters
            total_processed = 0
            added_count = 0
            updated_count = 0

            # Process the file in batches
            batch_size = 1000
            batch = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) < 3:
                    continue  # Skip rows with insufficient data
                
                id, ho_ten, dia_chi, ma_vung = row[:4] if len(row) > 3 else (*row[:3], None)
                
                # Clean and validate data
                id = str(id).strip() if id else None
                ho_ten = ILLEGAL_CHARACTERS_RE.sub('', str(ho_ten).strip()) if ho_ten else None
                dia_chi = ILLEGAL_CHARACTERS_RE.sub('', str(dia_chi).strip()) if dia_chi else None
                ma_vung = ILLEGAL_CHARACTERS_RE.sub('', str(ma_vung).strip()) if ma_vung else None

                if not id or not ho_ten:
                    continue  # Skip rows with missing essential data

                batch.append((id, ho_ten, dia_chi, ma_vung))
                
                if len(batch) >= batch_size:
                    process_batch(batch)
                    total_processed += len(batch)
                    batch = []

                if total_processed % 10000 == 0:
                    app.logger.info(f"Processed {total_processed} records")

            # Process any remaining records
            if batch:
                process_batch(batch)
                total_processed += len(batch)

            db.session.commit()
            app.logger.info(f"Hoàn thành import dữ liệu. Tổng số xử lý: {total_processed}")
        
        os.remove(filepath)
        return jsonify({
            'message': 'Import thành công',
            'total_processed': total_processed,
            'added': added_count,
            'updated': updated_count
        }), 200

    except Exception as e:
        app.logger.error(f"Error during import: {str(e)}")
        return jsonify({'error': f'Lỗi khi import: {str(e)}'}), 500

def process_batch(batch):
    for id, ho_ten, dia_chi, ma_vung in batch:
        existing_participant = db.session.get(Participant, id)
        if existing_participant:
            existing_participant.ho_ten = ho_ten
            existing_participant.dia_chi = dia_chi
            existing_participant.ma_vung = ma_vung
        else:
            new_participant = Participant(id=id, ho_ten=ho_ten, dia_chi=dia_chi, ma_vung=ma_vung)
            db.session.add(new_participant)
    db.session.flush()

@app.route('/api/check-data', methods=['GET'])
def check_data():
    awards = Award.query.all()
    participants = Participant.query.all()
    results = Result.query.all()
    return jsonify({
        'awards': [{'id': a.id, 'ten_giai': a.ten_giai, 'da_duoc_phat': a.da_duoc_phat} for a in awards],
        'participants': [{'id': p.id, 'ho_ten': p.ho_ten, 'dia_chi': p.dia_chi} for p in participants],
        'results': [{
            'participant_id': r.participant_id,
            'award_id': r.award_id,
            'thoi_gian': r.thoi_gian.isoformat()
        } for r in results]
    })

@app.route('/api/clear-participants', methods=['POST'])
def clear_participants():
    try:
        # Xóa tất cả kết quả trước
        Result.query.delete()
        # Sau đó xóa tất cả người tham gia
        Participant.query.delete()
        db.session.commit()
        return jsonify({'message': 'Đã xóa tất cả dữ liệu người tham gia và kết quả'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Lỗi khi xóa dữ liệu: {str(e)}'}), 500

@app.route('/api/participants', methods=['GET'])
def get_participants():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('limit', 10, type=int)
    participants = Participant.query.paginate(page=page, per_page=per_page)
    
    return jsonify({
        'participants': [{
            'id': participant.id,
            'ho_ten': participant.ho_ten,
            'dia_chi': participant.dia_chi,
            'ma_vung': participant.ma_vung
        } for participant in participants.items],
        'total': participants.total,
        'pages': participants.pages,
        'current_page': participants.page
    })

@app.route('/api/quay-so-manual', methods=['POST'])
def quay_so_manual():
    data = request.json
    award_id = data.get('awardId')
    input_number = data.get('inputNumber')
    
    if not award_id or input_number is None:
        return jsonify({'error': 'Thiếu awardId hoặc inputNumber'}), 400
    
    award = db.session.get(Award, award_id)
    if not award:
        return jsonify({'error': 'Giải thưởng không tồn tại'}), 404

    # Tìm người tham gia với số đã nhập
    participant = Participant.query.filter(func.substr(Participant.id, -len(str(input_number))) == str(input_number)).first()

    if not participant:
        return jsonify({'error': 'Không có mã số dự thưởng này'}), 404

    # Kiểm tra xem người này đã trúng bất kỳ giải nào chưa
    existing_result = Result.query.filter_by(participant_id=participant.id).first()
    if existing_result:
        return jsonify({'error': 'Đại lý đã trúng giải khác'}), 400

    # Tạo kết quả mới
    result = Result(participant_id=participant.id, award_id=award.id, thoi_gian=datetime.utcnow())
    db.session.add(result)
    db.session.commit()

    return jsonify({
        'message': 'Quay số thành công',
        'participant': {
            'id': participant.id,
            'ho_ten': participant.ho_ten,
            'dia_chi': participant.dia_chi
        },
        'award': {
            'id': award.id,
            'ten_giai': award.ten_giai,
            'gia_tri': award.gia_tri
        },
        'thoi_gian': result.thoi_gian.isoformat()
    }), 200

@app.route('/api/resultslive', methods=['GET'])
def get_results_live_view():
    page = request.args.get('page', 1, type=int)
    limit = request.args.get('limit', 10, type=int)
    award_id = request.args.get('award', '')

    query = Result.query.join(Participant).join(Award)

    if award_id:
        query = query.filter(Award.id == award_id)

    total_results = query.count()
    paginated_results = query.order_by(Result.thoi_gian).paginate(page=page, per_page=limit, error_out=False)

    return jsonify({
        'results': [{
            'id': result.id,
            'row_number': total_results - ((page - 1) * limit + index),
            'participant': {
                'id': result.participant.id,
                'ho_ten': result.participant.ho_ten,
                'dia_chi': result.participant.dia_chi,
                'ma_vung': result.participant.ma_vung
            },
            'award': {
                'id': result.award.id,
                'ten_giai': result.award.ten_giai,
                'gia_tri': result.award.gia_tri
            },
            'thoi_gian': result.thoi_gian.isoformat()
        } for index, result in enumerate(paginated_results.items)],
        'total': paginated_results.total,
        'pages': paginated_results.pages,
        'current_page': paginated_results.page
    })

@app.route('/api/quay-so/check', methods=['POST'])
def check_winner():
    data = request.get_json()
    award_id = data.get('award_id') or data.get('awardId')

    if not award_id:
        return jsonify({'error': 'Thiếu thông tin giải thưởng'}), 400

    try:
        # Lấy một người chơi ngẫu nhiên chưa trúng giải
        potential_winner = db.session.query(Participant)\
            .outerjoin(Result, Participant.id == Result.participant_id)\
            .filter(Result.id == None)\
            .order_by(func.random())\
            .first()

        if not potential_winner:
            return jsonify({'error': 'Không còn người chơi nào để quay số'}), 400

        # Kiểm tra xem người chơi đã trúng giải nào chưa
        existing_win = db.session.query(Result)\
            .join(Participant)\
            .filter(
                or_(
                    Participant.id == potential_winner.id,
                    and_(
                        Participant.ho_ten == potential_winner.ho_ten,
                        Participant.dia_chi == potential_winner.dia_chi
                    )
                )
            ).first()

        if existing_win:
            # Nếu đã trúng, trả về thông tin người chơi và giải đã trúng
            award = db.session.query(Award).filter_by(id=existing_win.award_id).first()
            return jsonify({
                'error': f'Đại lý đã trúng {award.ten_giai}',
                'participant': {
                    'id': potential_winner.id,
                    'ho_ten': potential_winner.ho_ten,
                    'dia_chi': potential_winner.dia_chi
                },
                'existing_award': {
                    'id': award.id,
                    'ten_giai': award.ten_giai
                }
            }), 409  # Conflict status code

        return jsonify({
            'participant': {
                'id': potential_winner.id,
                'ho_ten': potential_winner.ho_ten,
                'dia_chi': potential_winner.dia_chi
            }
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/quay-so/confirm', methods=['POST'])
def confirm_winner():
    data = request.get_json()
    award_id = data.get('award_id') or data.get('awardId')
    participant_id = data.get('participant_id')

    if not award_id or not participant_id:
        return jsonify({'error': 'Thiếu thông tin giải thưởng hoặc người tham gia'}), 400

    try:
        # Kiểm tra lại lần nữa trước khi lưu
        existing_win = db.session.query(Result)\
            .join(Participant)\
            .filter(
                or_(
                    Participant.id == participant_id,
                    and_(
                        Participant.ho_ten == Participant.query.get(participant_id).ho_ten,
                        Participant.dia_chi == Participant.query.get(participant_id).dia_chi
                    )
                )
            ).first()

        if existing_win:
            award = db.session.query(Award).filter_by(id=existing_win.award_id).first()
            return jsonify({
                'error': f'Đại lý đã trúng {award.ten_giai}',
                'participant': {
                    'id': participant_id,
                    'ho_ten': Participant.query.get(participant_id).ho_ten,
                    'dia_chi': Participant.query.get(participant_id).dia_chi
                },
                'existing_award': {
                    'id': award.id,
                    'ten_giai': award.ten_giai
                }
            }), 409

        # Tạo kết quả mới và lưu vào DB
        new_result = Result(
            participant_id=participant_id,
            award_id=award_id,
            thoi_gian=datetime.utcnow()
        )
        db.session.add(new_result)
        db.session.commit()

        return jsonify({
            'participant': {
                'id': participant_id,
                'ho_ten': Participant.query.get(participant_id).ho_ten,
                'dia_chi': Participant.query.get(participant_id).dia_chi
            },
            'award': {
                'id': award_id,
                'ten_giai': db.session.query(Award).filter_by(id=award_id).first().ten_giai
            }
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# Thêm middleware để xử lý CORS
@app.after_request
def after_request(response):
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    return response

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(host='0.0.0.0', port=5001)